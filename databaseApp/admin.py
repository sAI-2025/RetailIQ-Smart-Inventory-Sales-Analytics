from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.utils import timezone
import csv
import openpyxl
from openpyxl.utils import get_column_letter

from .models import OrderDetails


# 📤 Export as CSV
@admin.action(description="Download selected as CSV")
def export_as_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=order_details.csv'
    writer = csv.writer(response)

    fields = [field.name for field in OrderDetails._meta.fields]
    writer.writerow(fields)

    for obj in queryset:
        writer.writerow([getattr(obj, field) for field in fields])
    
    return response


# 📥 Export as Excel
@admin.action(description="Download selected as Excel")
def export_as_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Order Details'

    fields = [field.name for field in OrderDetails._meta.fields]
    for col_num, field_name in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = field_name

    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}{row_num}'] = getattr(obj, field_name)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=order_details.xlsx'
    workbook.save(response)
    return response


# ❌ Delete ALL rows (ignores selection)
@admin.action(description="Delete ALL rows (ignores selection)")
def delete_all_rows(modeladmin, request, queryset):
    OrderDetails.objects.all().delete()
    modeladmin.message_user(request, "✅ All rows have been deleted.")
    return HttpResponseRedirect(request.get_full_path())


# 🧤 Soft Delete
@admin.action(description="🗑️ Soft Delete selected")
def soft_delete_selected(modeladmin, request, queryset):
    for obj in queryset:
        obj.soft_delete()
    modeladmin.message_user(request, f"🗑️ {queryset.count()} orders soft-deleted.")


# 🔁 Restore
@admin.action(description="♻️ Restore selected soft-deleted orders")
def restore_selected(modeladmin, request, queryset):
    restored_count = 0
    for obj in queryset:
        if obj.is_deleted:
            obj.restore()
            restored_count += 1
    modeladmin.message_user(request, f"♻️ {restored_count} orders restored.")


# ✅ Admin Configuration
@admin.register(OrderDetails)
class OrderDetailsAdmin(admin.ModelAdmin):
    list_display = (
        'order_id', 'payment_id', 'device_id',
        'ordered_vanilla', 'ordered_chocolate', 'ordered_strawberry',
        'delivered_vanilla', 'delivered_chocolate', 'delivered_strawberry',
        'is_payment_successful', 'is_all_items_delivered',
        'date', 'time', 'is_deleted'
    )

    actions = [
        export_as_csv, export_as_excel,
        delete_all_rows, soft_delete_selected,
        restore_selected
    ]

    search_fields = ('order_id', 'payment_id', 'device_id')

    list_filter = (
        'is_payment_successful',
        'is_all_items_delivered',
        'date',
        'device_id',  # Added to filter by device ID
        'deleted_at',  # Added to easily filter deleted entries
    )

    readonly_fields = ('date', 'time',)






#admin.py


# from django.contrib import admin
# from django.urls import path
# from django.utils.html import format_html
# from django.shortcuts import redirect
# from django.contrib import messages
# from django.urls import reverse

# from .models import ProductStock


# @admin.register(ProductStock)
# class ProductStockAdmin(admin.ModelAdmin):
#     list_display = ('product_id', 'vanilla', 'chocolate', 'strawberry', 'admin_buttons')

#     def get_urls(self):
#         """Custom URLs for refill and fetch actions."""
#         urls = super().get_urls()
#         custom = [
#             path('fetch/', self.admin_site.admin_view(self.fetch_from_file), name='fetch_from_file'),
#             path('refill/<int:product_id>/<str:flavor>/', self.admin_site.admin_view(self.refill_stock), name='refill_flavor'),
#         ]
#         return custom + urls

#     def admin_buttons(self, obj):
#         """Create Refill buttons for each product flavor."""
#         html = ''
#         for flavor in ['vanilla', 'chocolate', 'strawberry']:
#             url = f'./refill/{obj.product_id}/{flavor}/'
#             html += f'<a class="button" href="{url}">Refill {flavor}</a>&nbsp;'
#         return format_html(html)

#     admin_buttons.short_description = "Refill Controls"

#     def fetch_from_file(self, request):
#         """Fetches the data from machine_state.py and updates the DB."""
#         # Importing inside the function to avoid circular import
#         from .utils import load_machine_state

#         data = load_machine_state()
#         for product_id, flavors in data.items():
#             obj, _ = ProductStock.objects.get_or_create(product_id=product_id)
#             for flavor, qty in flavors.items():
#                 setattr(obj, flavor, qty)
#             obj.save()

#         messages.success(request, "Fetched values from machine_state.py and updated database.")
#         changelist_url = reverse('admin:databaseApp_productstock_changelist')
#         return redirect(changelist_url)

#     def refill_stock(self, request, product_id, flavor):
#         """Refills the product's flavor stock by 150 units."""
#         # Importing inside the function to avoid circular import
#         from .utils import load_machine_state, write_machine_state

#         data = load_machine_state()
#         if product_id not in data:
#             messages.error(request, f"Product ID {product_id} not found.")
#             changelist_url = reverse('admin:databaseApp_productstock_changelist')
#             return redirect(changelist_url)

#         # Update stock in machine_state.py
#         new_value = data[product_id].get(flavor, 0) + 150
#         data[product_id][flavor] = new_value
#         write_machine_state(data)

#         # Update stock in the database
#         obj, _ = ProductStock.objects.get_or_create(product_id=product_id)
#         setattr(obj, flavor, new_value)
#         obj.save()

#         messages.success(request, f"{flavor.capitalize()} stock for product {product_id} refilled to {new_value}.")
#         changelist_url = reverse('admin:databaseApp_productstock_changelist')
#         return redirect(changelist_url)

from django.contrib import admin
from django.urls import path
from django.utils.html import format_html
from django.shortcuts import redirect
from django.contrib import messages
from django.urls import reverse

from .models import ProductStock


@admin.register(ProductStock)
class ProductStockAdmin(admin.ModelAdmin):
    list_display = ('product_id', 'vanilla_stock', 'chocolate_stock', 'strawberry_stock', 'admin_buttons')

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('refill/<int:product_id>/<str:flavor>/', self.admin_site.admin_view(self.refill_stock), name='refill_flavor'),
        ]
        return custom + urls

    def get_live_stock(self, product_id):
        """Fetch stock from machine_state.py using string keys."""
        from .machine_state import Product_live_stock_info
        return Product_live_stock_info.get(str(product_id), {})

    def vanilla_stock(self, obj):
        return self.get_live_stock(obj.product_id).get('vanilla', 'N/A')

    def chocolate_stock(self, obj):
        return self.get_live_stock(obj.product_id).get('chocolate', 'N/A')

    def strawberry_stock(self, obj):
        return self.get_live_stock(obj.product_id).get('strawberry', 'N/A')

    vanilla_stock.short_description = 'Vanilla'
    chocolate_stock.short_description = 'Chocolate'
    strawberry_stock.short_description = 'Strawberry'

    def admin_buttons(self, obj):
        """Render refill buttons for each flavor."""
        html = ''
        for flavor in ['vanilla', 'chocolate', 'strawberry']:
            url = f'./refill/{obj.product_id}/{flavor}/'
            html += f'<a class="button" href="{url}">Refill {flavor}</a>&nbsp;'
        return format_html(html)

    admin_buttons.short_description = "Refill Controls"

    def refill_stock(self, request, product_id, flavor):
        """Refill flavor by 150 units in machine_state.py."""
        from .utils import load_machine_state, write_machine_state

        data = load_machine_state()
        key = str(product_id)

        if key not in data:
            messages.error(request, f"Product ID {product_id} not found in machine state.")
            return redirect(reverse('admin:databaseApp_productstock_changelist'))

        current = data[key].get(flavor, 0)
        data[key][flavor] = current + 150
        write_machine_state(data)

        messages.success(request, f"{flavor.capitalize()} stock for product {product_id} refilled to {data[key][flavor]}.")
        return redirect(reverse('admin:databaseApp_productstock_changelist'))
